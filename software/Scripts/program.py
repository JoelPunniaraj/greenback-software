# Written By Joel Punniaraj 
# Last Updated : 4/9/2024

import os
import locale
import requests
import mysql.connector

import yfinance as yf

from prettytable import PrettyTable as pt
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter

# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.worksheet.datavalidation import DataValidation

locale.setlocale(locale.LC_ALL, '')

conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="datasprint",
    database="greenback"
)

c = conn.cursor()

c.execute('''CREATE TABLE IF NOT EXISTS portfolio
             (ticker VARCHAR(10), shares INT, cost_basis DECIMAL(10, 2))''')

def portfolio():
    try:
        main_menu()
    finally:
        conn.close()

def add_position(ticker, shares, cost_basis):
    c.execute("INSERT INTO portfolio VALUES (%s, %s, %s)", (ticker, shares, cost_basis))
    conn.commit()

def get_live_price(ticker):
    try:
        stock = yf.Ticker(ticker)
        return round(stock.history(period="1d").iloc[-1]['Close'], 2)
    except Exception as e:
        print(f"Error getting live price for {ticker}: {e}")
        return None

def is_valid_ticker(ticker):
    try:
        yf.Ticker(ticker).info
        return True
    except Exception:
        return False

def calculate_portfolio_value():
    total_value = 0
    c.execute("SELECT * FROM portfolio")
    rows = c.fetchall()
    for row in rows:
        ticker, shares, cost_basis = row
        live_price = get_live_price(ticker)
        if live_price is not None:
            position_value = shares * live_price
            total_value += position_value
    return total_value

def print_portfolio_table():

    sort_table = input("Do you want to sort the table first? (yes/no): ").lower()

    if sort_table == 'yes':
        while True:
            print("Sort Options:")
            print("- Sort by Value")
            print("- Sort by Profit")
            print("- Exit")
            print()
            sort_choice = input("Enter your sort choice: ").lower()
            
            if sort_choice == 'value':
                sort_key = lambda row: (row[1] * get_live_price(row[0])) if get_live_price(row[0]) is not None else 0
                reverse_sort = True
                break
            elif sort_choice == 'profit':
                sort_key = lambda row: row[6]
                reverse_sort = True
                break
            elif sort_choice == 'exit':
                return
            else:
                print("Invalid choice. Please enter a valid sort option.")
    else:
        sort_key = None
        reverse_sort = None

    table = pt(["Ticker", "Shares", "Cost Basis", "Initial Investment", "Current Price", "Position Value", "Profit/Loss", "% Change"])
    
    table.align["Ticker"] = "l"  
    table.align["Shares"] = "r"  
    table.align["Cost Basis"] = "r"  
    table.align["Initial Investment"] = "r" 
    table.align["Current Price"] = "r"  
    table.align["Position Value"] = "r" 
    table.align["Profit/Loss"] = "r" 
    table.align["% Change"] = "r" 
    
    c.execute("SELECT * FROM portfolio")
    rows = c.fetchall()
    
    if sort_key:
        sorted_rows = sorted(rows, key=sort_key, reverse=reverse_sort)
    else:
        sorted_rows = rows
        
    for row in sorted_rows:
        ticker, shares, cost_basis = row
        if ticker in ['BRKB', 'BRK.B']:
            continue
        live_price = get_live_price(ticker)
        if live_price is not None:
            position_value = shares * live_price
            initial_investment = shares * cost_basis
            profit_loss = position_value - initial_investment
            percent_change = ((live_price - cost_basis) / cost_basis) * 100 if cost_basis != 0 else 0
            formatted_cost_basis = locale.currency(cost_basis, grouping=True)
            formatted_live_price = locale.currency(live_price, grouping=True)
            formatted_position_value = locale.currency(position_value, grouping=True)
            formatted_initial_investment = locale.currency(initial_investment, grouping=True)
            formatted_profit_loss = locale.currency(profit_loss, grouping=True)
            formatted_percent_change = f"{percent_change:.2f}%"
            table.add_row([ticker, shares, formatted_cost_basis, formatted_initial_investment, 
            formatted_live_price, formatted_position_value, formatted_profit_loss, 
            formatted_percent_change])

    print(table)

    total_value = calculate_portfolio_value()
    formatted_total_value = locale.currency(total_value, grouping=True)
    print(f"\nTotal Value: {formatted_total_value}")
    print(f"\nTotal Profit: {formatted_total_value}")

def remove_position_menu():
    ticker = input("Enter the ticker symbol to remove: ").upper()
    c.execute("SELECT * FROM portfolio WHERE ticker=%s", (ticker,))
    position = c.fetchone()
    if position is None:
        print("Position not found.")
    else:
        confirm = input(f"Do you want to remove {ticker} from the portfolio? (yes/no): ").lower()
        if confirm == 'yes':
            c.execute("DELETE FROM portfolio WHERE ticker=%s", (ticker,))
            conn.commit()
            print(f"Position {ticker} removed successfully.")
        else:
            print("Operation canceled.")

def add_position_menu():
    while True:
        ticker = input("Enter the ticker symbol: ").upper()
        if is_valid_ticker(ticker):
            break
        else:
            print("Invalid ticker symbol. Please enter a valid ticker.")

    shares = int(input("Enter the number of shares: "))
    cost_basis = float(input("Enter the cost basis per share: "))
    add_position(ticker, shares, cost_basis)
    print("Position added successfully.")

def edit_position_menu():
    ticker = input("Enter the ticker symbol to edit: ").upper()
    c.execute("SELECT * FROM portfolio WHERE ticker=%s", (ticker,))
    position = c.fetchone()
    if position is None:
        print("Position not found.")
    else:
        print("Current position:", position)
        new_shares = int(input("Enter the new number of shares (or leave empty to skip): ") or position[1])
        new_cost_basis = float(input("Enter the new cost basis per share (or leave empty to skip): ") or position[2])
        c.execute("UPDATE portfolio SET shares=%s, cost_basis=%s WHERE ticker=%s", (new_shares, new_cost_basis, ticker))
        conn.commit()
        print("Position updated successfully.")

def get_live_price(ticker):
    try:
        if ticker in ['BRKB', 'BRK.B']:
            return None 
        stock = yf.Ticker(ticker)
        history = stock.history(period="1d")
        if not history.empty:
            return round(history.iloc[-1]['Close'], 2)
        else:
            return None
    except Exception as e:
        print(f"Error getting live price for {ticker}: {e}")
        return None
    
def is_valid_ticker(ticker):
    try:
        info = yf.Ticker(ticker).info
        if info:
            return True
        else:
            print(f"Invalid ticker symbol: {ticker}")
            return False
    except Exception as e:
        print(f"Error checking ticker symbol {ticker}: {e}")
        return False

def main_menu():
    try:
        while True:
            print("\nOptions:")
            print("- 'Add' ")
            print("- 'View' ")
            print("- 'Edit' ")
            print("- 'Remove' ")
            print("-  Exit" )
            print()
            choice = input("Enter your choice: ").lower()
            if choice == 'add':
                add_position_menu()
            elif choice == 'view':
                print_portfolio_table()
            elif choice == 'edit':
                edit_position_menu()
            elif choice == 'remove':
                remove_position_menu()
            elif choice == 'exit':
                break
            else:
                print("Invalid choice. Please enter a valid option.")
    finally:
        conn.close()


income_workbook_path = r"C:\Users\joelp\greenback\model\access_model.xlsx"
income_sheet_name = "income-sheet"
income_mapping = {
    "Revenue": 4,
    "Revenue Growth (YoY)": 5,
    "Cost of Revenue": 6,
    "Gross Profit": 7,
    "Selling, General & Admin": 8,
    "Research & Development": 9,
    "Operating Expenses": 10,
    "Operating Income": 11,
    "Interest Expense / Income": 12,
    "Other Expense / Income": 13,
    "Pretax Income": 14,
    "Income Tax": 15,
    "Net Income": 16,
    "Net Income Growth": 17,
    "Shares Outstanding (Basic)": 18,
    "Shares Outstanding (Diluted)": 19,
    "Shares Change": 20,
    "EPS (Basic)": 21,
    "EPS (Diluted)": 22,
    "EPS Growth": 23,
    "Free Cash Flow": 24,
    "Free Cash Flow Per Share": 25,
    "Dividend Per Share": 26,
    "Dividend Growth": 27,
    "Gross Margin": 28,
    "Operating Margin": 29,
    "Profit Margin": 30,
    "Free Cash Flow Margin": 31,
    "Effective Tax Rate": 32,
    "EBITDA": 33,
    "EBITDA Margin": 34,
    "Depreciation & Amortization": 35,
    "EBIT": 36,
    "EBIT Margin": 37
}

balance_workbook_path = r"C:\Users\joelp\greenback\model\access_model.xlsx"
balance_sheet_name = "balance-sheet"
balance_mapping = {
    "Cash & Equivalents": 4,
    "Short-Term Investments": 5,
    "Cash & Cash Equivalents": 6,
    "Cash Growth": 7,
    "Receivables": 8,
    "Inventory": 9,
    "Other Current Assets": 10,
    "Total Current Assets": 11,
    "Property, Plant & Equipment": 12,
    "Long-Term Investments": 13,
    "Goodwill and Intangibles": 14,
    "Other Long-Term Assets": 15,
    "Total Long-Term Assets": 16,
    "Total Assets": 17,
    "Accounts Payable": 18,
    "Deferred Revenue": 19,
    "Current Debt": 20,
    "Other Current Liabilities": 21,
    "Total Current Liabilities": 22,
    "Long-Term Debt": 23,
    "Other Long-Term Liabilities": 24,
    "Total Long-Term Liabilities": 25,
    "Total Liabilities": 26,
    "Total Debt": 27,
    "Debt Growth": 28,
    "Retained Earnings": 29, 
    "Comprehensive Income": 30,
    "Shareholders' Equity": 31,
    "Net Cash / Debt": 32,
    "Net Cash / Debt Growth": 33,
    "Net Cash Per Share": 34,
    "Working Capital": 35,
    "Book Value Per Share": 36
}

cash_flow_workbook_path = r"C:\Users\joelp\greenback\model\access_model.xlsx"
cash_flow_sheet_name = "cash-flow-sheet"
cash_flow_mapping = {
    "Net Income": 4,
    "Depreciation & Amortization": 5,
    "Share-Based Compensation": 6,
    "Other Operating Activities": 7, 
    "Operating Cash Flow": 8, 
    "Operating Cash Flow Growth": 9, 
    "Capital Expenditures": 10, 
    "Acquisitions": 11, 
    "Change in Investments": 12, 
    "Other Investing Activities": 13, 
    "Investing Cash Flow": 14, 
    "Dividends Paid": 15,
    "Share Issuance / Repurchase": 16, 
    "Debt Issued / Paid": 17,
    "Other Financing Activities": 18,
    "Financing Cash Flow": 19, 
    "Exchange Rate Effect": 20,
    "Net Cash Flow": 21,
    "Free Cash Flow": 22,
    "Free Cash Flow Growth": 23,
    "Free Cash Flow Margin": 24, 
    "Free Cash Flow Per Share": 25
}

market_cap_workbook_path = r"C:\Users\joelp\greenback\model\access_model.xlsx"
market_cap_sheet_name = "stock-history"
market_cap_mapping = {
    "Market Capitalization": 4,
    "Market Cap Growth": 5,
    "Enterprise Value": 6,
}

'''
def data_window():
    global quarters 
    quarters  = int(input("\nEnter The Number Of Years: "))
    quarters  = quarters - 1
'''

def import_income(url, income_mapping, income_workbook_path, income_sheet_name):
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find("tbody")
        rows = table.find_all("tr")  

        if not os.path.exists(income_workbook_path):
            workbook = Workbook()
        else:
            workbook = load_workbook(income_workbook_path)

        if income_sheet_name not in workbook.sheetnames:
            workbook.create_sheet(income_sheet_name)
        sheet = workbook[income_sheet_name]

        for row in rows:
            cells = row.find_all("td")
            metric = cells[0].get_text(strip=True)
            if metric in income_mapping:
                row_num = income_mapping[metric]
                values = [cell.get_text(strip=True) for cell in cells[1:quarters]] 

                for i, value in enumerate(values):
                    cell = sheet.cell(row=row_num, column=i+2)
                    '''
                    if value.replace('.', '', 1).replace(',', '').isdigit():
                        value = float(value.replace(',', ''))
                    '''
                    cell.value = value
                    cell.alignment = Alignment(horizontal='right')
        workbook.save(income_workbook_path)
        print("Income Sheet Data from 'stockanalysis.com' Added!")

    else:
        print("Failed to Fetch Data from the Website!")

def import_balance(url, balance_mapping, balance_workbook_path, balance_sheet_name):
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find("tbody")
        rows = table.find_all("tr")   

        if not os.path.exists(balance_workbook_path):
            workbook = Workbook()
        else:
            workbook = load_workbook(balance_workbook_path)

        if balance_sheet_name not in workbook.sheetnames:
            workbook.create_sheet(balance_sheet_name)
        sheet = workbook[balance_sheet_name]

        for row in rows:
            cells = row.find_all("td")
            metric = cells[0].get_text(strip=True)
            if metric in balance_mapping:
                row_num = balance_mapping[metric]
                values = [cell.get_text(strip=True) for cell in cells[1:quarters]] 

                for i, value in enumerate(values):
                    cell = sheet.cell(row=row_num, column=i+2)
                    '''
                    if value.replace('.', '', 1).replace(',', '').isdigit():
                        value = float(value.replace(',', ''))
                    '''
                    cell.value = value
                    cell.alignment = Alignment(horizontal='right') 
        workbook.save(balance_workbook_path)
        print("Balance Sheet Data from 'stockanalysis.com' Added!")

    else:
        print("Failed to Fetch Data from the Website!")

def import_cash_flow(url, cash_flow_mapping, cash_flow_workbook_path, cash_flow_sheet_name):
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find("tbody")
        rows = table.find_all("tr")   

        if not os.path.exists(cash_flow_workbook_path):
            workbook = Workbook()
        else:
            workbook = load_workbook(cash_flow_workbook_path)

        if cash_flow_sheet_name not in workbook.sheetnames:
            workbook.create_sheet(cash_flow_sheet_name)
        sheet = workbook[cash_flow_sheet_name]

        for row in rows:
            cells = row.find_all("td")
            metric = cells[0].get_text(strip=True)
            if metric in cash_flow_mapping:
                row_num = cash_flow_mapping[metric]
                values = [cell.get_text(strip=True) for cell in cells[1:quarters]] 

                for i, value in enumerate(values):
                    cell = sheet.cell(row=row_num, column=i+2)
                    '''
                    if value.replace('.', '', 1).replace(',', '').isdigit():
                        value = float(value.replace(',', ''))
                    '''
                    cell.value = value
                    
                    cell.alignment = Alignment(horizontal='right') 
        workbook.save(cash_flow_workbook_path)
        print("Cash Flow Sheet Data from 'stockanalysis.com' Added!")

    else:
        print("Failed to Fetch Data from the Website!")

def import_marketcap(url, market_cap_mapping, market_cap_workbook_path, market_cap_sheet_name):
    response = requests.get(url)

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, "html.parser")
        table = soup.find("tbody")
        rows = table.find_all("tr")   

        if not os.path.exists(market_cap_workbook_path):
            workbook = Workbook()
        else:
            workbook = load_workbook(market_cap_workbook_path)

        if market_cap_sheet_name not in workbook.sheetnames:
            workbook.create_sheet(market_cap_sheet_name)
        sheet = workbook[market_cap_sheet_name]

        for row in rows:
            cells = row.find_all("td")
            metric = cells[0].get_text(strip=True)
            if metric in market_cap_mapping:
                row_num = market_cap_mapping[metric]
                values = [cell.get_text(strip=True) for cell in cells[1:quarters]] 

                for i, value in enumerate(values):
                    cell = sheet.cell(row=row_num, column=i+2)
                    '''
                    if value.replace('.', '', 1).replace(',', '').isdigit():
                        value = float(value.replace(',', ''))
                    '''
                    cell.value = value
                    cell.alignment = Alignment(horizontal='right') 
        workbook.save(market_cap_workbook_path)
        print("Stock History Sheet Data from 'stockanalysis.com' Added!")

    else:
        print("Failed to Fetch Data from the Website!")

def convert_text_to_numbers(workbook_path, sheet_name, mapping):
    workbook = load_workbook(workbook_path)
    sheet = workbook[sheet_name]

    for metric, column_number in mapping.items(): # metric 'not defined'

        for row_num in range(2, sheet.max_row + 1):
            cell = sheet[get_column_letter(column_number) + str(row_num)]
            if cell.value and isinstance(cell.value, str) and cell.value.replace('.', '', 1).isdigit():
                cell.value = float(cell.value)
    workbook.save(workbook_path)

'''
def import_ticker():
    workbook = load_workbook(workbook_path)
    workbook_path = "C:\\Users\\joelp\\greenback\\access_model.xlsx"
    load = load_workbook(workbook_path)
    sheet = load["5y-projections"]
    cell = sheet["AX1"]
    cell.value = ticker
    workbook.save(workbook_path)
    pass
'''

def start_model():
    global ticker
    ticker = input("\nTicker: ").lower()
    print("\nAccessing 'https://stockanalysis.com/' for Raw Data...")
    income_url = f"https://stockanalysis.com/stocks/{ticker}/financials/?p=quarterly"
    balance_url = f"https://stockanalysis.com/stocks/{ticker}/financials/balance-sheet/?p=quarterly"
    cash_flow_url = f"https://stockanalysis.com/stocks/{ticker}/financials/cash-flow-statement/?p=quarterly"
    market_cap_url = f"https://stockanalysis.com/stocks/{ticker}/financials/ratios/?p=quarterly"

    import_income(income_url, income_mapping, income_workbook_path, income_sheet_name)
    import_balance(balance_url, balance_mapping, balance_workbook_path, balance_sheet_name)
    import_cash_flow(cash_flow_url, cash_flow_mapping, cash_flow_workbook_path, cash_flow_sheet_name)
    import_marketcap(market_cap_url, market_cap_mapping, market_cap_workbook_path, market_cap_sheet_name)

    convert_text_to_numbers(income_workbook_path, income_sheet_name, income_mapping)
    convert_text_to_numbers(balance_workbook_path, balance_sheet_name, balance_mapping)
    convert_text_to_numbers(cash_flow_workbook_path, cash_flow_sheet_name, cash_flow_mapping)
    convert_text_to_numbers(market_cap_workbook_path, market_cap_sheet_name, market_cap_mapping)

'''
def two_year_projections():
    workbook = load_workbook(income_workbook_path)
    sheet = workbook["2y-projections"]
    cell = sheet["AY6"]
    cell.value = ticker
    workbook.save(income_workbook_path)
    os.startfile(market_cap_workbook_path)
'''

def five_year_projections():
    workbook = load_workbook(income_workbook_path)
    sheet = workbook["projections"]
    cell = sheet["AY6"]
    cell.value = ticker
    workbook.save(income_workbook_path)
    os.startfile(market_cap_workbook_path)

'''
def remove_data():
    pass
'''

def projections():
    global quarters 
    print("\nProjection Option ")
    print("2y Model")
    print("5y Model")
    choice = input("\nEnter Option: ").lower()
    if choice == "5y":
        quarters = 21
        start_model()
        five_year_projections()
    '''
    elif choice == "2y":
        quarters = 9
        start_model()
        two_year_projections()
    '''

def valuations():
    print("\nDiscounted Cash Flow Model Options: ")
    print("- Growth")
    print("- Revenue")
    print("- EBITDA")

    option = input("\nEnter Option: ").lower()
    if option == "growth":
        print("\n- 5Y Model")
        print("- 10Y Model")
        choice = input("\nYear: ").lower()
        
        if choice == "5y":
            dcf_workbook = r"C:\Users\joelp\greenback\model\dcf_5y.xlsx"
            os.startfile(dcf_workbook)
            start_model()
        elif choice == "10y":
            pass
    elif option == "revenue":
        pass
    elif option == "ebitda":
        pass

def program():
    print("\nD A T A S P R I N T ,  I N C .")
    print("\nEnter Option:")
    print("- Projections")
    print("- Valuations")
    print("- Portfolio")
    print()

    choice = input("Enter Your Choice: ").lower()
    if choice == 'projections':
        projections()
    elif choice == 'valuations':
        valuations()
    elif choice == 'portfolio':
        portfolio()
    else:
        print("Invalid Choice! Please Enter a Valid Option.")

program()

if __name__ == "__main__":
    main_menu()
conn.close()




